import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, timedelta
import json
import os
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

class LeaveCalendarApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Team Leave Calendar")
        self.root.geometry("1400x900")
        self.data_file = "team_leaves.json"
        self.leaves = self.load_leaves()
        self.current_month = datetime.now().month
        self.current_year = datetime.now().year
        self.selected_employee = ""
        self.public_holidays = self.load_public_holidays()
        self.setup_ui()
        self.refresh_views()

    def load_public_holidays(self):
        holidays = {'NZ': {}, 'AU': {}, 'VN': {}, 'IN': {}}
        for year in range(2024, 2029):
            holidays['NZ'][f'{year}-01-01'] = 'New Year'
            holidays['NZ'][f'{year}-02-06'] = 'Waitangi Day'
            holidays['NZ'][f'{year}-04-25'] = 'ANZAC Day'
            holidays['NZ'][f'{year}-12-25'] = 'Christmas'
            holidays['NZ'][f'{year}-12-26'] = 'Boxing Day'
            holidays['AU'][f'{year}-01-01'] = 'New Year'
            holidays['AU'][f'{year}-01-26'] = 'Australia Day'
            holidays['AU'][f'{year}-04-25'] = 'ANZAC Day'
            holidays['AU'][f'{year}-12-25'] = 'Christmas'
            holidays['AU'][f'{year}-12-26'] = 'Boxing Day'
            holidays['VN'][f'{year}-01-01'] = 'New Year'
            holidays['VN'][f'{year}-04-30'] = 'Liberation Day'
            holidays['VN'][f'{year}-05-01'] = 'Labour Day'
            holidays['VN'][f'{year}-09-02'] = 'National Day'
            holidays['IN'][f'{year}-01-26'] = 'Republic Day'
            holidays['IN'][f'{year}-08-15'] = 'Independence Day'
            holidays['IN'][f'{year}-10-02'] = 'Gandhi Jayanti'
            holidays['IN'][f'{year}-12-25'] = 'Christmas'
        return holidays

    def load_leaves(self):
        if os.path.exists(self.data_file):
            try:
                with open(self.data_file, 'r') as f:
                    return json.load(f)
            except:
                return []
        return []

    def save_leaves(self):
        with open(self.data_file, 'w') as f:
            json.dump(self.leaves, f)
    def calculate_working_days(self, start_date, end_date):
        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d")
        working_days = 0
        current = start
        while current <= end:
            day_of_week = current.weekday()
            date_str = current.strftime("%Y-%m-%d")
            is_weekend = day_of_week >= 5
            is_holiday = any(date_str in country_holidays for country_holidays in self.public_holidays.values())
            if not is_weekend and not is_holiday:
                working_days += 1
            current += timedelta(days=1)
        return working_days

    def get_holiday_for_date(self, date_str):
        for country, holidays in self.public_holidays.items():
            if date_str in holidays:
                return country, holidays[date_str]
        return None, None

    def update_employee_filter(self):
        unique_employees = sorted(set(leave["name"] for leave in self.leaves))
        self.employee_filter["values"] = ["All Employees"] + unique_employees

    def get_filtered_leaves(self):
        if not self.selected_employee:
            return self.leaves
        return [leave for leave in self.leaves if leave["name"] == self.selected_employee]

    def filter_by_employee(self, event=None):
        selection = self.employee_filter.get()
        self.selected_employee = "" if selection == "All Employees" else selection
        self.refresh_views()

    def setup_ui(self):
        main_container = ttk.Frame(self.root, padding="10")
        main_container.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_container.columnconfigure(1, weight=1)
        main_container.rowconfigure(2, weight=1)
        
        title = ttk.Label(main_container, text="Team Leave Calendar", font=("Arial", 20, "bold"))
        title.grid(row=0, column=0, columnspan=2, pady=10)
        
        left_panel = ttk.Frame(main_container)
        left_panel.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 10))
        
        form_frame = ttk.LabelFrame(left_panel, text="Add New Leave", padding="10")
        form_frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(form_frame, text="Team Member Name:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.name_entry = ttk.Entry(form_frame, width=25)
        self.name_entry.grid(row=0, column=1, pady=5)
        
        ttk.Label(form_frame, text="Leave Type:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.leave_type = ttk.Combobox(form_frame, values=["planned", "tentative"], state="readonly", width=22)
        self.leave_type.set("planned")
        self.leave_type.grid(row=1, column=1, pady=5)
        
        ttk.Label(form_frame, text="Start Date (YYYY-MM-DD):").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.start_date = ttk.Entry(form_frame, width=25)
        self.start_date.grid(row=2, column=1, pady=5)
        
        ttk.Label(form_frame, text="End Date (YYYY-MM-DD):").grid(row=3, column=0, sticky=tk.W, pady=5)
        self.end_date = ttk.Entry(form_frame, width=25)
        self.end_date.grid(row=3, column=1, pady=5)
        
        ttk.Label(form_frame, text="Reason/Notes:").grid(row=4, column=0, sticky=tk.W, pady=5)
        self.reason_text = tk.Text(form_frame, width=25, height=3)
        self.reason_text.grid(row=4, column=1, pady=5)
        
        ttk.Label(form_frame, text="Available during Christmas:").grid(row=5, column=0, sticky=tk.W, pady=5)
        self.christmas_check = tk.BooleanVar()
        ttk.Checkbutton(form_frame, variable=self.christmas_check).grid(row=5, column=1, sticky=tk.W, pady=5)
        
        btn_frame = ttk.Frame(form_frame)
        btn_frame.grid(row=6, column=0, columnspan=2, pady=10)
        ttk.Button(btn_frame, text="Add Leave", command=self.add_leave).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Clear Form", command=self.clear_form).pack(side=tk.LEFT, padx=5)
        
        right_panel = ttk.Frame(main_container)
        right_panel.grid(row=1, column=1, sticky=(tk.W, tk.E, tk.N, tk.S))
        right_panel.columnconfigure(0, weight=1)
        right_panel.rowconfigure(1, weight=1)
        
        controls_frame = ttk.Frame(right_panel)
        controls_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Button(controls_frame, text="Calendar View", command=self.show_calendar).pack(side=tk.LEFT, padx=5)
        ttk.Button(controls_frame, text="List View", command=self.show_list).pack(side=tk.LEFT, padx=5)
        
        ttk.Label(controls_frame, text="Filter Employee:").pack(side=tk.LEFT, padx=(20, 5))
        self.employee_filter = ttk.Combobox(controls_frame, state="readonly", width=20)
        self.employee_filter.pack(side=tk.LEFT, padx=5)
        self.employee_filter.bind("<<ComboboxSelected>>", self.filter_by_employee)
        
        ttk.Button(controls_frame, text="Export to Excel", command=self.export_to_excel).pack(side=tk.LEFT, padx=5)
        
        self.notebook = ttk.Notebook(right_panel)
        self.notebook.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.calendar_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.calendar_frame, text="Calendar View")
        
        self.list_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.list_frame, text="List View")
        
        self.current_view = "calendar"

    def refresh_views(self):
        self.update_employee_filter()
        if self.current_view == "calendar":
            self.show_calendar()
        else:
            self.show_list()

    def show_calendar(self):
        self.current_view = "calendar"
        self.notebook.select(0)
        for widget in self.calendar_frame.winfo_children():
            widget.destroy()
        
        header_frame = ttk.Frame(self.calendar_frame)
        header_frame.pack(fill=tk.X, pady=10)
        
        prev_btn = ttk.Button(header_frame, text="< Previous", command=self.prev_month)
        prev_btn.pack(side=tk.LEFT, padx=10)
        
        month_label = ttk.Label(header_frame, text=calendar.month_name[self.current_month] + " " + str(self.current_year), font=("Arial", 14, "bold"))
        month_label.pack(side=tk.LEFT, padx=20)
        
        next_btn = ttk.Button(header_frame, text="Next >", command=self.next_month)
        next_btn.pack(side=tk.LEFT, padx=10)
        
        if self.current_year >= 2028 and self.current_month >= 12:
            next_btn.config(state=tk.DISABLED)
        if self.current_year <= 2024 and self.current_month <= 1:
            prev_btn.config(state=tk.DISABLED)
        
        legend_frame = ttk.Frame(self.calendar_frame)
        legend_frame.pack(fill=tk.X, pady=5)
        
        for country, color in [("New Zealand", "#2196f3"), ("Australia", "#4caf50"), ("Vietnam", "#fbc02d"), ("India", "#e91e63")]:
            frame = ttk.Frame(legend_frame)
            frame.pack(side=tk.LEFT, padx=10)
            label = tk.Label(frame, bg=color, width=3, height=1)
            label.pack(side=tk.LEFT, padx=2)
            ttk.Label(frame, text=country).pack(side=tk.LEFT)
        
        canvas = tk.Canvas(self.calendar_frame)
        scrollbar = ttk.Scrollbar(self.calendar_frame, orient=tk.VERTICAL, command=canvas.yview)
        calendar_container = ttk.Frame(canvas)
        canvas.create_window((0, 0), window=calendar_container, anchor=tk.NW)
        
        days = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
        for i, day in enumerate(days):
            label = ttk.Label(calendar_container, text=day, font=("Arial", 10, "bold"))
            label.grid(row=0, column=i, ipadx=5, ipady=5, sticky=(tk.W, tk.E))
        
        filtered_leaves = self.get_filtered_leaves()
        first_day = datetime(self.current_year, self.current_month, 1).weekday()
        first_day = (first_day + 1) % 7
        days_in_month = calendar.monthrange(self.current_year, self.current_month)[1]
        row = 1
        col = 0
        
        for day in range(1, days_in_month + 1):
            date_str = f"{self.current_year}-{self.current_month:02d}-{day:02d}"
            current_date = datetime(self.current_year, self.current_month, day)
            cell_frame = ttk.Frame(calendar_container, relief=tk.SUNKEN, borderwidth=1)
            cell_frame.grid(row=row, column=col, sticky=(tk.W, tk.E, tk.N, tk.S), padx=1, pady=1)
            
            holiday_country, holiday_name = self.get_holiday_for_date(date_str)
            if holiday_country:
                colors = {"NZ": "#e3f2fd", "AU": "#e8f5e9", "VN": "#fff9c4", "IN": "#fce4ec"}
                if holiday_country in colors:
                    cell_frame.configure(style="Holiday.TFrame")
            
            date_label = ttk.Label(cell_frame, text=str(day))
            date_label.pack(side=tk.TOP, anchor=tk.NW)
            
            leaves_today = [leave for leave in filtered_leaves if leave["startDate"] <= date_str <= leave["endDate"]]
            for leave in leaves_today[:3]:
                leave_color = "#0c5460" if leave["leaveType"] == "planned" else "#856404"
                bg_color = "#d1ecf1" if leave["leaveType"] == "planned" else "#fff3cd"
                leave_label = tk.Label(cell_frame, text=leave["name"], bg=bg_color, fg=leave_color, font=("Arial", 8), wraplength=60)
                leave_label.pack(side=tk.TOP, fill=tk.X, padx=2, pady=1)
                leave_label.bind("<Button-1>", lambda e, l=leave: self.edit_leave_modal(l))
            
            col += 1
            if col > 6:
                col = 0
                row += 1
        
        for r in range(10):
            calendar_container.rowconfigure(r, weight=1)
        for c in range(7):
            calendar_container.columnconfigure(c, weight=1)
        
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.update_idletasks()
        canvas.config(scrollregion=canvas.bbox("all"))

    def show_list(self):
        self.current_view = "list"
        self.notebook.select(1)
        for widget in self.list_frame.winfo_children():
            widget.destroy()
        
        list_container = ttk.Frame(self.list_frame)
        list_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        columns = ("Name", "Type", "Start Date", "End Date", "Duration", "Christmas Available", "Reason")
        tree = ttk.Treeview(list_container, columns=columns, show="headings", height=20)
        
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=120)
        
        scrollbar = ttk.Scrollbar(list_container, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        filtered_leaves = self.get_filtered_leaves()
        for leave in sorted(filtered_leaves, key=lambda x: x["startDate"], reverse=True):
            duration = self.calculate_working_days(leave["startDate"], leave["endDate"])
            christmas = "Yes" if leave.get("christmasAvailable", False) else "No"
            tree.insert("", tk.END, values=(leave["name"], leave["leaveType"].capitalize(), leave["startDate"], leave["endDate"], f"{duration} working days", christmas, leave.get("reason", "-")))
        
        tree.bind("<Double-1>", lambda e: self.edit_selected_leave(tree))

    def prev_month(self):
        if self.current_month > 1:
            self.current_month -= 1
        else:
            self.current_month = 12
            self.current_year -= 1
        if self.current_year < 2024:
            self.current_year = 2024
            self.current_month = 1
        self.show_calendar()

    def next_month(self):
        if self.current_month < 12:
            self.current_month += 1
        else:
            self.current_month = 1
            self.current_year += 1
        if self.current_year > 2028 or (self.current_year == 2028 and self.current_month > 12):
            self.current_year = 2028
            self.current_month = 12
        self.show_calendar()

    def add_leave(self):
        try:
            name = self.name_entry.get().strip()
            leave_type = self.leave_type.get()
            start_date = self.start_date.get().strip()
            end_date = self.end_date.get().strip()
            reason = self.reason_text.get(1.0, tk.END).strip()
            christmas_available = self.christmas_check.get()
            
            if not name or not start_date or not end_date:
                messagebox.showwarning("Validation Error", "Please fill in all required fields")
                return
            
            if start_date > end_date:
                messagebox.showwarning("Validation Error", "End date must be after start date")
                return
            
            leave = {"id": datetime.now().timestamp(), "name": name, "leaveType": leave_type, "startDate": start_date, "endDate": end_date, "reason": reason, "christmasAvailable": christmas_available}
            self.leaves.append(leave)
            self.save_leaves()
            self.clear_form()
            self.refresh_views()
            messagebox.showinfo("Success", "Leave entry added successfully")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    def clear_form(self):
        self.name_entry.delete(0, tk.END)
        self.leave_type.set("planned")
        self.start_date.delete(0, tk.END)
        self.end_date.delete(0, tk.END)
        self.reason_text.delete(1.0, tk.END)
        self.christmas_check.set(False)

    def edit_selected_leave(self, tree):
        selection = tree.selection()
        if not selection:
            return
        item = tree.item(selection[0])
        name = item["values"][0]
        leave = next((l for l in self.leaves if l["name"] == name), None)
        if leave:
            self.edit_leave_modal(leave)

    def edit_leave_modal(self, leave):
        modal = tk.Toplevel(self.root)
        modal.title("Edit Leave")
        modal.geometry("400x300")
        modal.transient(self.root)
        modal.grab_set()
        
        ttk.Label(modal, text="Edit Leave Entry", font=("Arial", 12, "bold")).pack(pady=10)
        frame = ttk.Frame(modal, padding="10")
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Name:").grid(row=0, column=0, sticky=tk.W, pady=5)
        name_entry = ttk.Entry(frame, width=30)
        name_entry.insert(0, leave["name"])
        name_entry.grid(row=0, column=1, pady=5)
        
        ttk.Label(frame, text="Type:").grid(row=1, column=0, sticky=tk.W, pady=5)
        type_combo = ttk.Combobox(frame, values=["planned", "tentative"], state="readonly", width=27)
        type_combo.set(leave["leaveType"])
        type_combo.grid(row=1, column=1, pady=5)
        
        ttk.Label(frame, text="Start Date:").grid(row=2, column=0, sticky=tk.W, pady=5)
        start_entry = ttk.Entry(frame, width=30)
        start_entry.insert(0, leave["startDate"])
        start_entry.grid(row=2, column=1, pady=5)
        
        ttk.Label(frame, text="End Date:").grid(row=3, column=0, sticky=tk.W, pady=5)
        end_entry = ttk.Entry(frame, width=30)
        end_entry.insert(0, leave["endDate"])
        end_entry.grid(row=3, column=1, pady=5)
        
        button_frame = ttk.Frame(frame)
        button_frame.grid(row=4, column=0, columnspan=2, pady=20)
        
        def save_edit():
            try:
                if messagebox.askyesno("Confirm Delete", "This will update the leave entry. Continue?"):
                    leave["name"] = name_entry.get()
                    leave["leaveType"] = type_combo.get()
                    leave["startDate"] = start_entry.get()
                    leave["endDate"] = end_entry.get()
                    self.save_leaves()
                    self.refresh_views()
                    modal.destroy()
            except Exception as e:
                messagebox.showerror("Error", str(e))
        
        ttk.Button(button_frame, text="Save", command=save_edit).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=modal.destroy).pack(side=tk.LEFT, padx=5)

    def export_to_excel(self):
        try:
            from tkinter import filedialog
            filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], initialfile=f"Team_Leaves_{datetime.now().strftime('%Y-%m-%d')}.xlsx")
            if not filename:
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Leave List"
            
            headers = ["Name", "Leave Type", "Start Date", "End Date", "Duration (Working Days)", "Available during Christmas Shutdown", "Reason/Notes"]
            ws.append(headers)
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
            
            for leave in self.leaves:
                duration = self.calculate_working_days(leave["startDate"], leave["endDate"])
                christmas = "Yes" if leave.get("christmasAvailable", False) else "No"
                ws.append([leave["name"], leave["leaveType"].capitalize(), leave["startDate"], leave["endDate"], duration, christmas, leave.get("reason", "")])
            
            from openpyxl.utils import get_column_letter
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
            
            wb.save(filename)
            messagebox.showinfo("Success", f"Data exported successfully to {filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = LeaveCalendarApp(root)
    root.mainloop()
